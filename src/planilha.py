"""
Planilha orçamentária da contratação.

Substitui o campo único de valor estimado por uma tabela de itens
(código, descrição, unidade, quantidade, valor unitário). O valor total
de cada item e o valor global (soma = estimativa da contratação) são
derivados automaticamente. Os itens podem ser digitados na tabela ou
importados de um arquivo XLSX.
"""

import io
import re
import unicodedata

# Chaves numéricas/estruturais usadas no cálculo (código, descrição etc.)
CAMPOS_ITEM = ["codigo", "descricao", "unidade", "quantidade", "valor_unitario"]

# Coluna opcional fixa para a fonte do preço (geralmente um link/URL)
CAMPO_FONTE = "fonte"

# Colunas derivadas (não editáveis) — não entram no editor
CAMPOS_DERIVADOS = {"valor_total"}

# Sinônimos de cabeçalho aceitos na importação de XLSX (sem acento, minúsculo)
SINONIMOS = {
    "codigo": ["codigo", "cod", "item", "n", "no", "num", "numero"],
    "descricao": ["descricao", "especificacao", "discriminacao", "objeto",
                  "descricao do item", "especificacoes", "produto", "servico"],
    "unidade": ["unidade", "und", "un", "unid", "medida", "unidade de medida", "um"],
    "quantidade": ["quantidade", "qtd", "qtde", "quant", "qte", "qtd."],
    "valor_unitario": ["valor unitario", "vlr unitario", "vlr unit", "preco unitario",
                       "valor unit", "unitario", "vl unitario", "preco unit", "p unit"],
    "fonte": ["fonte", "link", "url", "referencia", "origem", "endereco",
              "fonte do preco", "fonte do valor", "site", "pesquisa"],
}

# Rótulos amigáveis para as colunas do editor
ROTULOS = {
    "codigo": "Código",
    "descricao": "Descrição",
    "unidade": "Unidade",
    "quantidade": "Quantidade",
    "valor_unitario": "Valor Unitário (R$)",
    "valor_total": "Valor Total (R$)",
    "fonte": "Fonte / Link",
}

_RE_URL = re.compile(r"^\s*(https?://|www\.)\S+\s*$", re.IGNORECASE)


def eh_url(valor) -> bool:
    return bool(_RE_URL.match(str(valor or "")))


def normalizar_url(valor) -> str:
    """Garante esquema http(s) para o link ficar clicável."""
    url = str(valor or "").strip()
    if url.lower().startswith("www."):
        return "https://" + url
    return url


def para_link_markdown(valor) -> str:
    """URL -> '[link](url)' (compacto e clicável); demais valores inalterados."""
    if eh_url(valor):
        return f"[link]({normalizar_url(valor)})"
    return str(valor or "")


# ---------------------------------------------------------------------------
# Limpeza de texto (descrições vindas de PDF costumam ter espaços espúrios
# no meio de palavras: "plás tica", "docu mentos", "d?água")
# ---------------------------------------------------------------------------
# Fragmentos que NÃO são palavras isoladas em português: quando aparecem
# soltos, quase sempre são o final de uma palavra quebrada por um espaço.
# Comparados SEM acento (via _core). Propositalmente omitidos os que colidem
# com palavras reais: "do/da/ha/ao/as" (palavras), "sao"→são, "ida", "cidade",
# "idade", "grafica"→gráfica, "menta" etc. — juntá-los corromperia texto.
_FRAGMENTOS = {
    "tica", "tico", "ticas", "ticos", "oplastica",
    "mento", "mentos",
    "ado", "ada", "ados", "adas",
    "avel", "aveis",
    "cao", "coes",          # capta ção/ções (risco de "cão" é desprezível aqui)
    "enio",
    "dade", "dades",
    "encia", "encias", "ancia", "ancias",
    "essidade", "bilidade", "tividade",
}

_RE_APOSTROFO = re.compile(r"(?<=[A-Za-zÀ-ÿ])\?(?=[A-Za-zÀ-ÿ])")
_RE_ESPACO_PONT = re.compile(r"\s+([,.;:!?)])")
_RE_ESPACOS = re.compile(r"\s{2,}")


def _core(token: str) -> str:
    """Token sem acentos, minúsculo e sem pontuação de borda (p/ comparar)."""
    t = token.strip(".,;:!?)(-–—\"'")
    t = unicodedata.normalize("NFKD", t)
    t = "".join(c for c in t if not unicodedata.combining(c))
    return t.lower()


def limpar_texto(valor) -> str:
    """
    Corrige artefatos comuns de texto copiado de PDF:
      - '?' entre letras vira apóstrofo (d?água -> d'água);
      - espaço antes de pontuação e espaços duplicados;
      - junta uma palavra quebrada quando o pedaço seguinte é claramente um
        fragmento de sufixo (plás tica -> plástica, docu mentos -> documentos).
    Conservador: só junta quando o 2º pedaço não é uma palavra real, para não
    colar texto legítimo (ex.: "de expediente" permanece intacto).
    """
    s = str(valor or "")
    if not s.strip():
        return s
    s = s.replace("\xa0", " ").replace("​", "")
    s = _RE_APOSTROFO.sub("'", s)
    s = _RE_ESPACO_PONT.sub(r"\1", s)
    s = _RE_ESPACOS.sub(" ", s).strip()

    tokens = s.split(" ")
    saida: list[str] = []
    i = 0
    while i < len(tokens):
        atual = tokens[i]
        proximo = tokens[i + 1] if i + 1 < len(tokens) else ""
        if (
            atual and atual.isalpha() and 2 <= len(atual) <= 12
            and proximo and _core(proximo) in _FRAGMENTOS
        ):
            saida.append(atual + proximo)  # mantém a pontuação do fragmento
            i += 2
            continue
        saida.append(atual)
        i += 1
    return " ".join(saida)


def linha_vazia() -> dict:
    return {"codigo": "", "descricao": "", "unidade": "", "quantidade": 0.0,
            "valor_unitario": 0.0}


def linhas_iniciais(n: int = 3) -> list[dict]:
    return [linha_vazia() for _ in range(n)]


def _num(valor) -> float:
    """Converte para float, aceitando moeda BR ('R$ 1.234,56') e strings."""
    if valor is None or valor == "":
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip()
    for lixo in ("r$", "R$", " ", "\xa0"):
        texto = texto.replace(lixo, "")
    # padrão BR: ponto de milhar, vírgula decimal
    if "," in texto and "." in texto:
        texto = texto.replace(".", "").replace(",", ".")
    elif "," in texto:
        texto = texto.replace(",", ".")
    try:
        return float(texto)
    except ValueError:
        return 0.0


def _normalizar(texto) -> str:
    """minúsculo, sem acento, sem pontuação de borda — para casar cabeçalhos."""
    t = unicodedata.normalize("NFKD", str(texto or ""))
    t = "".join(c for c in t if not unicodedata.combining(c))
    return t.strip().lower().strip(".:")


class ErroPlanilha(Exception):
    """Erro amigável ao importar a planilha de um arquivo."""


# Raízes (substring, sem acento) para casar cabeçalhos escritos por extenso,
# ex.: "Descrição dos Serviços", "Especificação do Objeto", "Preço Unitário".
# 'valor_unitario' antes de qualquer coisa que contenha só 'valor'/'preco'
# para não confundir com "Valor Total".
_RAIZES = [
    ("valor_unitario", ("unitar", "vlr unit", "vl unit", "p unit", "preco unit",
                         "valor unit", "custo unit")),
    ("quantidade", ("quant", "qtd", "qtde", "qte")),
    ("descricao", ("descric", "especific", "discrimin", "objeto", "produto",
                   "servico", "item ", "material", "insumo")),
    ("codigo", ("codig", "cod ", "cod.", "sku", "referencia interna")),
    ("unidade", ("unidade", "und", "unid", "medida")),
    ("fonte", ("fonte", "link", "url", "origem", "endereco", "site",
               "pesquisa", "cotacao")),
]


def _campo_do_cabecalho(celula) -> str | None:
    """
    Mapeia um texto de cabeçalho para um campo do item. Tenta, nesta ordem:
    igualdade com um sinônimo, palavra inteira igual a um sinônimo e, por
    fim, raiz por substring — assim 'Descrição dos Serviços' vira 'descricao'.
    """
    norm = _normalizar(celula)
    if not norm:
        return None
    # 1) igualdade exata com um sinônimo
    for campo, syns in SINONIMOS.items():
        if norm in syns:
            return campo
    # 2) alguma palavra do cabeçalho é exatamente um sinônimo (>= 3 letras)
    palavras = norm.split()
    for campo, syns in SINONIMOS.items():
        for s in syns:
            if len(s) >= 3 and s in palavras:
                return campo
    # 3) raiz por substring (nomes escritos por extenso)
    for campo, raizes in _RAIZES:
        if any(r in norm for r in raizes):
            return campo
    return None


def _mapear_linha(linha) -> dict:
    """Devolve {coluna: campo} reconhecidos numa possível linha de cabeçalho."""
    mapa: dict[int, str] = {}
    for col, celula in enumerate(linha):
        campo = _campo_do_cabecalho(celula)
        if campo and campo not in mapa.values():
            mapa[col] = campo
    return mapa


def importar_de_xlsx(dados: bytes) -> list[dict]:
    """
    Lê um arquivo XLSX e devolve a lista de itens. Detecta a linha de
    cabeçalho pelos nomes das colunas (aceita acentos, nomes por extenso e
    variações); se não houver cabeçalho reconhecível, assume a ordem código,
    descrição, unidade, quantidade, valor unitário. Ignora linhas sem descrição.
    """
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(dados), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ErroPlanilha(
            f"Não foi possível ler o arquivo XLSX: {exc}. "
            "Envie um arquivo Excel (.xlsx) válido."
        ) from exc

    # varre TODAS as abas com dados (a 1ª aba pode ser capa/instruções)
    melhor = None  # (nº de campos, aba, idx_cabecalho, mapa, linhas)
    for ws in wb.worksheets:
        linhas = [list(r) for r in ws.iter_rows(values_only=True)]
        if not any(any(c not in (None, "") for c in ln) for ln in linhas):
            continue
        for i, linha in enumerate(linhas[:25]):
            mapa = _mapear_linha(linha)
            # cabeçalho válido: reconhece a DESCRIÇÃO (coluna essencial) e
            # pelo menos mais uma coluna (quantidade, valor, código...).
            if "descricao" in mapa.values() and len(set(mapa.values())) >= 2:
                pontos = len(set(mapa.values()))
                if melhor is None or pontos > melhor[0]:
                    melhor = (pontos, i, mapa, linhas)
                break

    itens: list[dict] = []
    cabecalho_visto = ""
    if melhor is not None:
        _, idx_cabecalho, mapa, linhas = melhor
        cabecalho = linhas[idx_cabecalho]
        cabecalho_visto = " | ".join(
            str(c).strip() for c in cabecalho if c not in (None, "")
        )
        # colunas não reconhecidas são preservadas com o rótulo original
        # (mas ignoramos "Valor Total"/"Total": esse valor é recalculado)
        extras = {
            col: str(cabecalho[col]).strip()
            for col in range(len(cabecalho))
            if col not in mapa and cabecalho[col] not in (None, "")
            and "total" not in _normalizar(cabecalho[col])
        }
        for linha in linhas[idx_cabecalho + 1:]:
            item = {c: "" for c in CAMPOS_ITEM}
            for col, campo in mapa.items():
                if col < len(linha):
                    item[campo] = linha[col]
            for col, rotulo in extras.items():
                if col < len(linha) and linha[col] not in (None, ""):
                    item[rotulo] = linha[col]
            _acrescentar(itens, item)
    else:
        # sem cabeçalho reconhecível: assume ordem posicional na 1ª aba com dados
        for ws in wb.worksheets:
            linhas = [list(r) for r in ws.iter_rows(values_only=True)]
            if not any(any(c not in (None, "") for c in ln) for ln in linhas):
                continue
            for linha in linhas:
                if not any(c not in (None, "") for c in linha):
                    continue
                item = dict(zip(CAMPOS_ITEM, list(linha) + [""] * len(CAMPOS_ITEM)))
                _acrescentar(itens, item)
            break

    if not itens:
        dica = (
            f' O cabeçalho lido foi: "{cabecalho_visto}".' if cabecalho_visto
            else " Não foi encontrada uma linha de cabeçalho."
        )
        raise ErroPlanilha(
            "Nenhum item reconhecido. A planilha precisa de uma coluna de "
            "descrição (ou especificação/objeto) e, de preferência, quantidade "
            "e valor unitário." + dica +
            " Dica: baixe o modelo abaixo e cole os seus dados nele."
        )
    return itens


def modelo_xlsx() -> bytes:
    """Gera um arquivo XLSX-modelo com o cabeçalho esperado e um exemplo."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Planilha Orçamentária"
    ws.append(["Código", "Descrição", "Unidade", "Quantidade",
               "Valor Unitário", "Fonte / Link"])
    ws.append(["001", "Notebook corporativo i5 16GB", "un", 100, 4500.00,
               "https://www.exemplo.com/notebook-i5"])
    ws.append(["002", "Monitor 24 polegadas", "un", 100, 900.00,
               "https://www.exemplo.com/monitor-24"])
    for col, larg in zip("ABCDEF", (10, 42, 12, 14, 16, 34)):
        ws.column_dimensions[col].width = larg
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _acrescentar(itens: list[dict], item: dict) -> None:
    """Normaliza os tipos e adiciona o item se tiver descrição."""
    descricao = str(item.get("descricao") or "").strip()
    if not descricao:
        return
    registro = {
        "codigo": str(item.get("codigo") or "").strip(),
        "descricao": limpar_texto(descricao),
        "unidade": str(item.get("unidade") or "").strip(),
        "quantidade": _num(item.get("quantidade")),
        "valor_unitario": _num(item.get("valor_unitario")),
    }
    # preserva a fonte e quaisquer colunas extras (texto)
    for chave, valor in item.items():
        if chave in registro or chave in CAMPOS_DERIVADOS:
            continue
        if valor not in (None, ""):
            texto = str(valor).strip()
            # não mexe em URLs (fonte/link); limpa demais textos
            registro[chave] = texto if eh_url(texto) else limpar_texto(texto)
    itens.append(registro)


def colunas_extra(itens: list[dict]) -> list[str]:
    """
    Colunas além das fixas (código..valor total) presentes em algum item,
    em ordem estável: 'fonte' primeiro, depois as demais na ordem de
    aparição. Usadas no editor, no prompt e na exportação.
    """
    fixas = set(CAMPOS_ITEM) | CAMPOS_DERIVADOS
    extras: list[str] = []
    for item in itens or []:
        for chave in item:
            if chave not in fixas and chave not in extras:
                extras.append(chave)
    if CAMPO_FONTE in extras:
        extras.remove(CAMPO_FONTE)
        extras.insert(0, CAMPO_FONTE)
    return extras


def item_valido(item: dict) -> bool:
    """Considera preenchido o item com descrição e algum valor/quantidade."""
    return bool((item.get("descricao") or "").strip()) and (
        _num(item.get("quantidade")) > 0 or _num(item.get("valor_unitario")) > 0
    )


def calcular(itens: list[dict]) -> tuple[list[dict], float]:
    """
    Filtra itens válidos, calcula valor_total de cada um (quantidade ×
    valor unitário) e o valor global (soma). Retorna (itens, valor_global).
    """
    resultado: list[dict] = []
    global_ = 0.0
    for item in itens or []:
        if not item_valido(item):
            continue
        qtd = _num(item.get("quantidade"))
        unit = _num(item.get("valor_unitario"))
        total = round(qtd * unit, 2)
        global_ += total
        registro = {
            "codigo": str(item.get("codigo") or "").strip(),
            "descricao": limpar_texto(str(item.get("descricao") or "").strip()),
            "unidade": str(item.get("unidade") or "").strip(),
            "quantidade": qtd,
            "valor_unitario": unit,
            "valor_total": total,
        }
        # preserva fonte e colunas extras (texto); limpa texto, mas não URLs
        for chave, valor in item.items():
            if chave in registro or chave in CAMPOS_DERIVADOS:
                continue
            texto = "" if valor is None else str(valor).strip()
            if texto:
                registro[chave] = texto if eh_url(texto) else limpar_texto(texto)
        resultado.append(registro)
    return resultado, round(global_, 2)


def formatar_moeda(valor) -> str:
    """R$ 1.234.567,89 (padrão brasileiro)."""
    v = _num(valor)
    return "R$ " + f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _rotulo_coluna(chave: str) -> str:
    """Rótulo de exibição de uma coluna extra (usa ROTULOS ou o próprio nome)."""
    return ROTULOS.get(chave, chave)


def para_markdown(itens: list[dict], valor_global: float,
                  incluir_global: bool = True) -> str:
    """
    Tabela Markdown da planilha, com colunas extras (ex.: Fonte/Link) e o
    valor global na última linha. Links são compactados para '[link](url)'
    — clicáveis e enxutos nos documentos exportados. Com incluir_global=False
    omite a linha do VALOR GLOBAL (usado em amostras de contexto).
    """
    if not itens:
        return "(planilha não informada)"
    extras = colunas_extra(itens)
    cabecalho = ["Código", "Descrição", "Unidade", "Quantidade",
                 "Valor Unitário", "Valor Total"] + [_rotulo_coluna(e) for e in extras]
    linhas = [
        "| " + " | ".join(cabecalho) + " |",
        "|" + "---|" * len(cabecalho),
    ]
    for it in itens:
        qtd = f"{it['quantidade']:g}"
        # Item ainda não passado por calcular() não traz 'valor_total'; o
        # produto é derivado aqui em vez de sair R$ 0,00 na coluna. A
        # tabela é determinística: nenhuma célula pode ser um vazio
        # formatado como dinheiro.
        total = it.get("valor_total")
        if total is None:
            total = round(_num(it.get("quantidade"))
                          * _num(it.get("valor_unitario")), 2)
        celulas = [
            it.get("codigo") or "-", it.get("descricao") or "",
            it.get("unidade") or "-", qtd,
            formatar_moeda(it.get("valor_unitario")),
            formatar_moeda(total),
        ]
        for e in extras:
            celulas.append(para_link_markdown(it.get(e, "")) or "-")
        linhas.append("| " + " | ".join(str(c) for c in celulas) + " |")
    if incluir_global:
        fim = ["", "", "", "", "**VALOR GLOBAL**",
               f"**{formatar_moeda(valor_global)}**"] + [""] * len(extras)
        linhas.append("| " + " | ".join(fim) + " |")
    return "\n".join(linhas)


# A tabela de itens é DETERMINÍSTICA: nasce da planilha do processo e é
# injetada por código, em QUALQUER tamanho. O modelo nunca escreve linha
# de item — nem quando são poucos. (Antes, até 12 itens o prompt mandava
# a IA reproduzir a planilha inteira, e acima disso ainda levava 6 linhas
# REAIS como "amostra ilustrativa": material pronto para ser copiado. Foi
# essa amostra que reapareceu como tabela parcial no edital do caso
# auditado — 53 dos 210 códigos, em 3 fragmentos soltos.)
MARCADOR_TABELA = "[[TABELA_ITENS]]"

# Mantido por compatibilidade: nenhum caminho pede mais a tabela à IA,
# então a contagem não decide nada.
LIMITE_ITENS_INLINE = 12


# ---------------------------------------------------------------------------
# Resumo SEMÂNTICO da planilha
#
# Retirar a planilha do prompt resolveu a cópia, mas criou outro problema:
# sem saber O QUE se compra, o modelo escreve DFD, ETP e TR genéricos —
# não há como justificar a necessidade, definir requisitos ou fixar
# critérios de recebimento para "210 itens" abstratos.
#
# A saída é dar COMPOSIÇÃO FUNCIONAL sem dar CONTEÚDO REPRODUZÍVEL: em
# que famílias os itens se agrupam e com que peso, sem um único código,
# preço, quantidade, link ou descrição literal. O modelo entende o objeto;
# não consegue remontar a planilha.
# ---------------------------------------------------------------------------
FAMILIAS_ITENS: dict[str, tuple[str, ...]] = {
    "Papelaria e expediente": (
        "papel", "caneta", "lapis", "borracha", "caderno", "envelope",
        "grampeador", "grampo", "clipe", "cola", "tesoura", "regua",
        "marcador", "pincel atomico", "corretivo", "apontador",
        "almofada para carimbo", "carimbo", "percevejo", "alfinete"),
    "Arquivo e organização de documentos": (
        "pasta", "arquivo morto", "caixa arquivo", "fichario",
        "porta-documento", "classificador", "divisoria", "etiqueta"),
    "Impressão e suprimentos de informática": (
        "toner", "cartucho", "tinta para impressora", "pen drive",
        "mouse", "teclado", "cabo", "midia", "dvd", "cd"),
    "Limpeza e higienização": (
        "detergente", "desinfetante", "agua sanitaria", "sabao",
        "papel higienico", "papel toalha", "alcool", "vassoura", "rodo",
        "pano de chao", "saco de lixo", "luva de latex"),
    "Copa e cozinha": (
        "copo descartavel", "cafe", "acucar", "adocante", "filtro de cafe",
        "guardanapo", "colher descartavel"),
    "Mobiliário e utensílios": (
        "cadeira", "mesa", "armario", "estante", "quadro branco",
        "gaveteiro", "suporte"),
}
_FAMILIA_OUTROS = "Outros materiais do objeto"

# Abaixo deste percentual a família não é citada isoladamente: entra no
# agregado, para que o resumo descreva a composição e não a lista.
_PISO_FAMILIA_PCT = 3.0


def _familia_do_item(descricao: str) -> str:
    texto = _normalizar(descricao or "")
    for familia, termos in FAMILIAS_ITENS.items():
        if any(t in texto for t in termos):
            return familia
    return _FAMILIA_OUTROS


def composicao_por_familia(itens: list[dict]) -> list[tuple[str, int, float]]:
    """
    (família, nº de itens, % do total de itens), da maior para a menor.

    Conta ITENS, nunca valores: percentual financeiro permitiria estimar
    preços por engenharia reversa, e o modelo não precisa disso para
    entender a composição funcional do objeto.
    """
    if not itens:
        return []
    contagem: dict[str, int] = {}
    for item in itens:
        familia = _familia_do_item(item.get("descricao"))
        contagem[familia] = contagem.get(familia, 0) + 1
    total = len(itens)
    return sorted(
        ((f, n, round(100.0 * n / total, 1)) for f, n in contagem.items()),
        key=lambda linha: (-linha[1], linha[0]))


def resumo_semantico(itens: list[dict]) -> str:
    """
    Composição funcional do objeto em prosa, sem nada reproduzível.

    NÃO contém: código, descrição literal, quantidade, preço, unidade
    monetária, link, nem linha de tabela. Contém: em que famílias os
    itens se agrupam e o peso relativo de cada uma.
    """
    composicao = composicao_por_familia(itens)
    if not composicao:
        return ""
    principais = [c for c in composicao if c[2] >= _PISO_FAMILIA_PCT]
    resto = [c for c in composicao if c[2] < _PISO_FAMILIA_PCT]
    partes = [f"{familia} ({pct:g}% dos itens)"
              for familia, _, pct in principais]
    if resto:
        pct_resto = round(sum(c[2] for c in resto), 1)
        partes.append(f"outras famílias de menor expressão ({pct_resto:g}%)")
    return (
        "COMPOSIÇÃO FUNCIONAL DO OBJETO (para você compreender o que se "
        "contrata; NÃO reproduza esta análise como lista): "
        + "; ".join(partes) + ". "
        "Use isso para fundamentar a necessidade, os requisitos, o modelo "
        "de execução, a fiscalização e os critérios de recebimento de "
        "forma pertinente a ESTAS famílias — e não em termos genéricos."
    )


def resumo_para_prompt(itens: list[dict], valor_global: float) -> str:
    """
    O que a IA recebe sobre a planilha: SOMENTE estatística e o marcador.

    Nenhum código, descrição, quantidade, preço ou link real entra no
    prompt — não há o que copiar. A tabela completa é inserida depois,
    por código, no lugar da marca.
    """
    n = len(itens)
    unidades = sorted({(it.get("unidade") or "").strip()
                       for it in itens if (it.get("unidade") or "").strip()})
    valores = [it.get("valor_unitario") or 0 for it in itens]
    faixa = (f" Preços unitários entre {formatar_moeda(min(valores))} e "
             f"{formatar_moeda(max(valores))}." if valores else "")
    return (
        f"A planilha orçamentária do processo possui {n} item(ns). "
        f"VALOR GLOBAL (estimativa total da contratação) = "
        f"{formatar_moeda(valor_global)}."
        + (f" Unidades de fornecimento: {', '.join(unidades[:12])}."
           if unidades else "")
        + faixa + "\n"
        f"PROIBIDO escrever a lista de itens, ainda que parcialmente: nada "
        f"de códigos, descrições, quantidades, preços, links ou linhas de "
        f"tabela — nem a partir do memorando ou dos anexos (se eles "
        f"contiverem a lista, ignore-a: a tabela oficial vem da planilha do "
        f"sistema). A TABELA COMPLETA, com todas as colunas e o valor "
        f"global, é inserida AUTOMATICAMENTE no lugar da marca "
        f"{MARCADOR_TABELA}.\n"
        f"Escreva o texto da cláusula de estimativa de valor (metodologia, "
        f"fundamento e conclusão) e coloque a marca {MARCADOR_TABELA} "
        f"EXATAMENTE UMA VEZ, SOZINHA em uma linha própria, dentro dessa "
        f"cláusula.\n"
        + resumo_semantico(itens)
    )


_RE_CABECALHO_ITENS = re.compile(
    r"^\s*\|\s*C[óo]digo\s*\|", re.IGNORECASE | re.MULTILINE)
_RE_SEPARADOR = re.compile(r"^\s*\|[\s:|-]+\|?\s*$")
_RE_LINHA_TABELA = re.compile(r"^\s*\|")
# linha do VALOR GLOBAL (rodapé da tabela, sem código na 1ª célula)
_RE_LINHA_GLOBAL = re.compile(r"VALOR\s+GLOBAL", re.IGNORECASE)

# Tabela sem o cabeçalho canônico ainda é lista de itens quando tem
# porte de planilha e a primeira coluna é sistematicamente um código.
_MINIMO_LINHAS_PLANILHA = 5
_PROPORCAO_CODIGOS = 0.8


def _celulas(linha: str) -> list[str]:
    return [c.strip() for c in linha.strip().strip("|").split("|")]


def _blocos_de_tabela(linhas: list[str]) -> list[tuple[int, int]]:
    """Intervalos [ini, fim) de linhas consecutivas de tabela Markdown."""
    blocos, ini = [], None
    for i, linha in enumerate(linhas):
        if _RE_LINHA_TABELA.match(linha):
            ini = i if ini is None else ini
            continue
        if ini is not None:
            blocos.append((ini, i))
            ini = None
    if ini is not None:
        blocos.append((ini, len(linhas)))
    return blocos


def _linhas_de_item(bloco: list[str]) -> list[str]:
    """Linhas de dados do bloco (sem cabeçalho, separador e VALOR GLOBAL)."""
    return [ln for ln in bloco
            if not _RE_SEPARADOR.match(ln)
            and not _RE_CABECALHO_ITENS.match(ln)
            and not _RE_LINHA_GLOBAL.search(ln)]


def _e_tabela_de_itens(bloco: list[str]) -> bool:
    """
    O bloco é a planilha orçamentária?

    Vale o cabeçalho canônico ("| Código | Descrição |", que a injeção
    escreve) OU o formato: muitas linhas cuja primeira célula é um código
    numérico. O segundo critério pega cópias do modelo com cabeçalho
    renomeado, sem confundir a matriz de riscos ou o cronograma — que têm
    poucas linhas e primeira coluna textual. O código NÃO é reconhecido
    por quantidade de dígitos: a planilha real mistura 3 e 6 dígitos.
    """
    if any(_RE_CABECALHO_ITENS.match(ln) for ln in bloco):
        return True
    dados = _linhas_de_item(bloco)
    if len(dados) < _MINIMO_LINHAS_PLANILHA:
        return False
    numericas = sum(1 for ln in dados
                    if (_celulas(ln) or [""])[0].isdigit())
    return numericas >= _PROPORCAO_CODIGOS * len(dados)


def linhas_de_itens_do_texto(texto: str) -> list[str]:
    """Todas as linhas de item das tabelas de planilha do documento."""
    linhas = (texto or "").splitlines()
    achadas: list[str] = []
    for ini, fim in _blocos_de_tabela(linhas):
        bloco = linhas[ini:fim]
        if _e_tabela_de_itens(bloco):
            achadas.extend(_linhas_de_item(bloco))
    return achadas


def remover_tabelas_da_ia(texto: str) -> tuple[str, int]:
    """
    Remove os blocos de tabela de itens que o MODELO escreveu.

    A planilha é dado determinístico do processo: a única tabela legítima
    é a injetada por código. Qualquer outra é cópia (da amostra antiga do
    prompt, do memorando ou de um anexo) — e vinha parcial, fora de ordem
    e sem conferência, como no edital auditado. Tabelas de prosa (riscos,
    cronograma) são preservadas. Devolve o texto limpo e quantas linhas
    foram removidas.
    """
    linhas = (texto or "").splitlines()
    remover: set[int] = set()
    removidas = 0
    for ini, fim in _blocos_de_tabela(linhas):
        if _e_tabela_de_itens(linhas[ini:fim]):
            remover.update(range(ini, fim))
            removidas += fim - ini
    if not remover:
        return (texto or "").strip(), 0
    saida = [ln for i, ln in enumerate(linhas) if i not in remover]
    return re.sub(r"\n{3,}", "\n\n", "\n".join(saida)).strip(), removidas


def injetar_tabela(texto: str, itens_brutos: list[dict] | None) -> str:
    """
    Deixa no documento UMA tabela de itens: a da planilha do processo.

    Ordem das operações (todas determinísticas):
      1. tabelas escritas pelo modelo são REMOVIDAS — sempre, mesmo com
         o marcador presente. Sem isto, o documento saía com a cópia
         parcial da IA e a tabela oficial (defeito real: DFD com a
         planilha duplicada, edital com 53 de 210 códigos);
      2. a tabela oficial entra no lugar do marcador; se o modelo
         esqueceu o marcador, entra ao final, em qualquer tamanho —
         nenhum documento pode ficar sem a planilha;
      3. marcadores extras são apagados (uma planilha de centenas de
         linhas duplicada inviabiliza o documento);
      4. a tabela ocupa um BLOCO próprio: colada na prosa, o conversor
         DOCX promove o primeiro item a cabeçalho e o repete em toda
         página.
    """
    itens, glob = calcular(itens_brutos or [])
    texto = texto or ""
    if not itens:
        return texto.replace(MARCADOR_TABELA, "").strip()

    texto, _ = remover_tabelas_da_ia(texto)
    tabela = para_markdown(itens, glob)
    if MARCADOR_TABELA in texto:
        antes, _, depois = texto.partition(MARCADOR_TABELA)
        depois = depois.replace(MARCADOR_TABELA, "")
        return (
            antes.rstrip() + "\n\n" + tabela + "\n\n" + depois.lstrip()
        ).strip()
    return (texto.rstrip() + "\n\n" + tabela).strip()


# ---------------------------------------------------------------------------
# Conferência da tabela emitida CONTRA A FONTE (validação determinística)
# ---------------------------------------------------------------------------
def conferir_tabela(texto: str, itens_brutos: list[dict] | None) -> list[str]:
    """
    Divergências entre a tabela do documento e a planilha do processo.

    Confere o CONJUNTO INTEGRAL: uma única tabela, todos os códigos
    (nenhum faltando, nenhum estranho, nenhum repetido), quantidade,
    unidade e preço unitário de cada item, e o valor global. Devolve
    lista vazia quando a tabela reproduz a fonte exatamente.
    """
    itens, glob = calcular(itens_brutos or [])
    if not itens:
        return []
    texto = texto or ""
    problemas: list[str] = []

    linhas_doc = texto.splitlines()
    tabelas = [(i, f) for i, f in _blocos_de_tabela(linhas_doc)
               if _e_tabela_de_itens(linhas_doc[i:f])]
    if not tabelas:
        return [f"tabela de itens ausente do documento "
                f"({len(itens)} item(ns) na planilha do processo)"]
    if len(tabelas) > 1:
        problemas.append(
            f"tabela de itens aparece {len(tabelas)} vezes no documento "
            "(a planilha oficial é única)")

    linhas = linhas_de_itens_do_texto(texto)
    no_doc: dict[str, list[str]] = {}
    for ln in linhas:
        cel = _celulas(ln)
        no_doc.setdefault(cel[0], cel)

    esperados = {str(it.get("codigo") or "").strip(): it for it in itens}
    faltando = [c for c in esperados if c not in no_doc]
    estranhos = [c for c in no_doc if c not in esperados]
    repetidos = len(linhas) - len(no_doc)
    if faltando:
        problemas.append(
            f"{len(faltando)} item(ns) da planilha não constam da tabela "
            f"(ex.: {', '.join(sorted(faltando)[:5])})")
    if estranhos:
        problemas.append(
            f"{len(estranhos)} item(ns) na tabela não existem na planilha "
            f"(ex.: {', '.join(sorted(estranhos)[:5])})")
    if repetidos:
        problemas.append(f"{repetidos} linha(s) de item repetida(s)")

    divergentes = []
    for codigo, item in esperados.items():
        cel = no_doc.get(codigo)
        if not cel or len(cel) < 6:
            continue
        _, _, unidade, qtd, unitario, total = cel[:6]
        if unidade != (item.get("unidade") or "-"):
            divergentes.append(f"{codigo} (unidade)")
        elif qtd != f"{item['quantidade']:g}":
            divergentes.append(f"{codigo} (quantidade)")
        elif unitario != formatar_moeda(item.get("valor_unitario")):
            divergentes.append(f"{codigo} (valor unitário)")
        elif total != formatar_moeda(item.get("valor_total")):
            # Quantidade e unitário conferem, mas o total escrito na linha
            # não é o produto dos dois. Sem esta conferência, um documento
            # com aritmética errada passa linha a linha e só destoa na
            # soma — se destoar.
            divergentes.append(f"{codigo} (valor total)")
    if divergentes:
        problemas.append(
            f"{len(divergentes)} item(ns) com valor divergente da planilha "
            f"(ex.: {', '.join(divergentes[:5])})")

    if formatar_moeda(glob) not in texto:
        problemas.append(
            f"valor global da planilha ({formatar_moeda(glob)}) não consta "
            "do documento")
    return problemas
