"""
Relatórios da pesquisa de preços (§31, §32, §34).

O relatório completo é a **memória do ato**: é o que um auditor lê meses
depois para decidir se o preço estimado se sustenta. Por isso ele tem
duas propriedades que governam este módulo inteiro:

**1. Ele contém o que foi descartado.** O §31 pede, em itens separados,
"todas as referências selecionadas" (13) e "referências desconsideradas
e motivo" (14). Um relatório que mostrasse só a cesta seria uma defesa,
não uma memória — e a pergunta que a auditoria faz é justamente por que
os outros preços não entraram.

**2. Ele não inventa.** Onde o dado não existe, o relatório escreve que
não existe. "(não informado)" é informação; um campo em branco é dúvida
sobre se ninguém preencheu ou se o sistema perdeu.

O formato de saída é **Markdown**, e isso é uma decisão de arquitetura: o
§33 proíbe um segundo pipeline de PDF, e o projeto já tem
`export.gerar_pdf(titulo, markdown, branding)` — DOCX estilizado →
LibreOffice → PDF, com as larguras de tabela e o gate de geometria já
provados. Este módulo produz o conteúdo; a conversão é do motor que já
existe.

Nada aqui toca banco, sessão ou rede.
"""

from __future__ import annotations

import hashlib
import io
import json
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation

from .. import planilha
from .estados import EstadoItem
from .perfil import obter as obter_perfil

# Versão do formato do relatório. Entra no identificador: mudar o layout
# muda o documento, e dois PDFs diferentes não podem carregar o mesmo
# identificador de versão.
FORMATO = "1"

AUSENTE = "(não informado)"

_ROTULO_STATUS = {
    "selected": "selecionada",
    "candidate": "candidata",
    "rejected": "excluída",
    "warning": "sinalizada",
    "manual_review": "abaixo do piso de comparabilidade",
}


def _decimal(valor) -> Decimal | None:
    if valor in (None, ""):
        return None
    try:
        return Decimal(str(valor))
    except (InvalidOperation, ValueError, TypeError):
        return None


def _moeda(valor) -> str:
    numero = _decimal(valor)
    return AUSENTE if numero is None else planilha.formatar_moeda(float(numero))


def _texto(valor) -> str:
    limpo = str(valor or "").strip()
    return limpo or AUSENTE


def _numero(valor) -> str:
    numero = _decimal(valor)
    if numero is None:
        return AUSENTE
    return f"{numero:g}"


def _data(valor) -> str:
    if not valor:
        return AUSENTE
    texto = str(valor)[:10]
    try:
        return date.fromisoformat(texto).strftime("%d/%m/%Y")
    except (ValueError, TypeError):
        return texto


# ---------------------------------------------------------------------------
# §34 — identificador da versão
# ---------------------------------------------------------------------------
def identificador_da_versao(pesquisa: dict, itens: list[dict],
                            referencias: dict[str, list[dict]]) -> str:
    """
    Impressão digital do RESULTADO — o item 22 do §31.

    Cobre exatamente o que define o resultado e nada de decorativo:
    cabeçalho da pesquisa, método e preço de cada item, e, de cada
    referência, o `raw_hash` (que já é a impressão da evidência como a
    fonte a entregou) com o status que ela acabou tendo.

    O que ele permite: pegar o PDF de hoje, refazer a projeção meses
    depois e provar que é o mesmo resultado — ou descobrir que alguém
    mexeu. É por isso que ele NÃO inclui data de geração: dois relatórios
    do mesmo resultado, emitidos em dias diferentes, têm de bater.
    """
    projecao = {
        "formato": FORMATO,
        "pesquisa": {
            "id": str(pesquisa.get("id") or ""),
            "versao": int(pesquisa.get("versao") or 1),
            "perfil": str(pesquisa.get("perfil_normativo") or ""),
            "data_base": str(pesquisa.get("data_base") or ""),
        },
        "itens": [],
    }
    for item in sorted(itens, key=lambda i: int(i.get("numero") or 0)):
        do_item = {
            "numero": int(item.get("numero") or 0),
            "descricao": str(item.get("descricao") or ""),
            "estado": str(item.get("estado") or ""),
            "metodo": str(item.get("metodo") or ""),
            "preco_estimado": str(item.get("preco_estimado") or ""),
            "referencias": sorted(
                (f"{r.get('raw_hash') or ''}:{r.get('status') or ''}"
                 for r in referencias.get(str(item.get("id")), [])),
            ),
        }
        projecao["itens"].append(do_item)
    canonico = json.dumps(projecao, sort_keys=True, ensure_ascii=False,
                          separators=(",", ":"), default=str)
    return hashlib.sha256(canonico.encode("utf-8")).hexdigest()


# ---------------------------------------------------------------------------
# §31 — relatório completo
# ---------------------------------------------------------------------------
def completo(pesquisa: dict, itens: list[dict],
             referencias: dict[str, list[dict]],
             *, emitido_em: datetime | None = None,
             processo: dict | None = None) -> str:
    """
    As 22 seções do §31, em Markdown, na ordem do prompt.

    A ordem não é estética: ela vai do CONTEXTO (quem, o quê, sob qual
    regra) para o MÉTODO (como se buscou, como se comparou, como se
    calculou) e só então para o RESULTADO. Um relatório que abrisse com o
    número convidaria a ler só o número.
    """
    momento = emitido_em or datetime.now(timezone.utc)
    perfil = obter_perfil(pesquisa.get("perfil_normativo"))
    concluidos = [i for i in itens
                  if str(i.get("estado")) == EstadoItem.COMPLETO.value]

    partes: list[str] = []
    ad = partes.append

    ad("# RELATÓRIO DE PESQUISA DE PREÇOS")
    ad("")
    ad("Lei nº 14.133/2021 — formação do preço estimado da contratação.")
    ad("")

    # 1–4
    ad("## 1. Identificação")
    ad("")
    ad(f"- **Pesquisa:** {_texto(pesquisa.get('nome'))}")
    ad(f"- **Identificador:** {_texto(pesquisa.get('id'))}")
    ad(f"- **Revisão:** {int(pesquisa.get('versao') or 1)}")
    if pesquisa.get("motivo_da_revisao"):
        ad(f"- **Motivo da revisão:** {_texto(pesquisa['motivo_da_revisao'])}")
    ad(f"- **Situação:** {_texto(pesquisa.get('estado'))}")
    ad(f"- **Data-base:** {_data(pesquisa.get('data_base'))}")
    ad("")

    ad("## 2. Objeto")
    ad("")
    ad(_texto(pesquisa.get("objeto")))
    ad("")

    ad("## 3. Processo")
    ad("")
    if pesquisa.get("processo_id"):
        ad(f"- **Processo vinculado:** {_texto(pesquisa.get('processo_id'))}")
        if processo:
            ad(f"- **Órgão:** {_texto(processo.get('orgao'))}")
            ad(f"- **Objeto do processo:** {_texto(processo.get('objeto'))}")
    else:
        ad("Pesquisa autônoma — ainda não vinculada a processo.")
    ad("")

    ad("## 4. Responsáveis")
    ad("")
    ad(f"- **Responsável indicado:** {_texto(pesquisa.get('responsavel'))}")
    ad(f"- **Local de referência:** {_texto(pesquisa.get('local_referencia'))}")
    ad("")

    # 5
    ad("## 5. Base normativa e perfil utilizado")
    ad("")
    ad(f"- **Perfil:** {perfil.nome}")
    ad(f"- **Base legal:** {perfil.base_legal}")
    ad(f"- **Mínimo de referências por item:** {perfil.minimo_referencias}")
    ad("- **Admite concluir com menos, mediante justificativa:** "
       + ("sim" if perfil.admite_menos_com_justificativa else "não"))
    if perfil.teto_da_mediana_em_sistema_oficial:
        ad("- **Teto da mediana:** aplicável — estimativa apoiada "
           "exclusivamente em sistema oficial de preços não supera a "
           "mediana da amostra.")
    for observacao in perfil.observacoes:
        ad(f"- {observacao}")
    ad("")

    # 6
    ad("## 6. Metodologia")
    ad("")
    ad("A pesquisa coleta referências em fontes oficiais, normaliza a "
       "unidade de fornecimento, avalia a comparabilidade de cada "
       "referência com o item da contratação, monta a cesta e forma o "
       "preço por método estatístico declarado.")
    ad("")
    ad("Nenhuma etapa preenche lacuna por estimativa: referência sem "
       "preço, sem unidade conversível ou sem comparabilidade suficiente "
       "fica fora da cesta, com o motivo registrado, e o item é "
       "declarado incompleto em vez de receber preço fabricado.")
    ad("")
    metodos = sorted({str(i.get("metodo") or "") for i in concluidos
                      if i.get("metodo")})
    ad(f"- **Métodos aplicados nesta pesquisa:** "
       f"{', '.join(metodos) if metodos else AUSENTE}")
    ad("")

    # 7
    ad("## 7. Fontes consultadas")
    ad("")
    ad(_tabela_de_fontes(referencias))
    ad("")

    # 8
    ad("## 8. Critérios de busca")
    ad("")
    filtros = pesquisa.get("filtros") or {}
    janela = filtros.get("janela_dias")
    ad(f"- **Janela temporal:** últimos {janela} dias contados da "
       f"data-base" if janela else "- **Janela temporal:** " + AUSENTE)
    ad(f"- **Recorte geográfico:** {_texto(filtros.get('uf'))}")
    ad("- **Código de catálogo (CATMAT/CATSER):** aceito e usado quando "
       "informado; **não é exigido** — a busca por descrição opera na "
       "ausência dele.")
    ad("")

    # 9
    ad("## 9. Critérios de comparabilidade")
    ad("")
    ad("A comparabilidade de cada referência é o produto de dois juízos "
       "independentes:")
    ad("")
    ad("- **identidade** — é o mesmo produto? (código de catálogo e "
       "semelhança de descrição);")
    ad("- **circunstâncias** — a contratação é comparável? (unidade, "
       "temporalidade, quantidade, geografia e condições).")
    ad("")
    ad("A identidade **multiplica** as circunstâncias em vez de disputar "
       "peso com elas: produto diferente zera a nota por melhor que "
       "seja o resto. Cada fator é registrado separadamente, com o peso "
       "declarado, e aparece na análise item a item.")
    ad("")

    # 10
    ad("## 10. Critérios estatísticos")
    ad("")
    ad("- **Método automático:** mediana quando a série é dispersa "
       "(coeficiente de variação acima de 0,25) e média quando é "
       "homogênea — a mediana resiste a valores extremos, e é com "
       "dispersão alta que os extremos distorcem a média;")
    ad("- **Sinalização de discrepantes:** intervalo interquartil "
       "(Tukey, fator 1,5) e desvio absoluto mediano (fator 3);")
    ad("- **Discrepante sinalizado NÃO é excluído automaticamente.** A "
       "exclusão é ato do revisor, com motivo registrado.")
    ad("")

    # 11
    ad("## 11. Quadro-resumo")
    ad("")
    ad(quadro_resumo(itens))
    ad("")

    # 12–17
    ad("## 12. Análise item a item")
    ad("")
    # Os itens 13 a 17 do §31 são POR ITEM — referências usadas,
    # referências descartadas, memória de cálculo, unitário e total.
    # Repeti-los como seções de topo os faria aparecer uma vez só, quando
    # existe um conjunto deles para cada item. Ficam dentro de cada
    # bloco, com a correspondência dita em voz alta para quem procura a
    # numeração do prompt.
    ad("Cada item abaixo traz, na ordem: as referências selecionadas "
       "(item 13 da estrutura), as desconsideradas com o motivo de cada "
       "exclusão (14), a memória de cálculo (15), o preço estimado "
       "unitário (16) e o valor total do item (17).")
    ad("")
    for item in sorted(itens, key=lambda i: int(i.get("numero") or 0)):
        for linha in _analise_do_item(item, referencias):
            ad(linha)

    # 18
    ad("## 18. Valor global")
    ad("")
    total = sum((_decimal(i.get("preco_total")) or Decimal("0"))
                for i in concluidos)
    ad(f"- **Itens concluídos:** {len(concluidos)} de {len(itens)}")
    ad(f"- **Valor global dos itens concluídos:** {_moeda(total)}")
    pendentes = [i for i in itens if i not in concluidos]
    if pendentes:
        numeros = ", ".join(str(int(i.get("numero") or 0))
                            for i in pendentes)
        ad(f"- **Itens sem preço formado:** {numeros}. O valor global "
           "acima **não** os inclui.")
    ad("")

    # 19
    ad("## 19. Alertas e ressalvas")
    ad("")
    for linha in _alertas(itens, referencias, perfil):
        ad(linha)
    ad("")

    # 20
    ad("## 20. Evidências e anexos")
    ad("")
    ad("Cada referência listada guarda, no banco desta pesquisa, o "
       "identificador oficial da contratação de origem e a impressão "
       "digital (SHA-256) do registro exatamente como a fonte o "
       "devolveu. A memória analítica completa acompanha este relatório "
       "em planilha (XLSX).")
    ad("")
    ad("A reprodutibilidade não depende de URL viva: o conteúdo "
       "coletado está preservado, e a impressão digital permite provar "
       "que é o mesmo registro.")
    ad("")

    # 21–22
    ad("## 21. Data e hora da pesquisa")
    ad("")
    ad(f"- **Relatório emitido em:** "
       f"{momento.strftime('%d/%m/%Y às %H:%M')} (UTC)")
    ad(f"- **Pesquisa criada em:** {_data(pesquisa.get('criado_em'))}")
    if pesquisa.get("aplicada_em"):
        ad(f"- **Aplicada ao processo em:** "
           f"{_data(pesquisa.get('aplicada_em'))}")
    ad("")

    ad("## 22. Identificador da versão")
    ad("")
    identificador = identificador_da_versao(pesquisa, itens, referencias)
    ad(f"`{identificador}`")
    ad("")
    ad("Impressão digital SHA-256 do resultado — cabeçalho da pesquisa, "
       "método e preço de cada item, e a evidência de cada referência "
       "com o status que ela recebeu. Não inclui a data de emissão: dois "
       "relatórios do mesmo resultado, emitidos em dias diferentes, "
       "produzem o mesmo identificador.")
    ad("")
    return "\n".join(partes)


def _tabela_de_fontes(referencias: dict[str, list[dict]]) -> str:
    """§31.7 — quais fontes entraram, e com quantas referências."""
    contagem: dict[tuple[str, str], int] = {}
    for lista in referencias.values():
        for linha in lista:
            chave = (str(linha.get("fonte_nome") or linha.get("fonte_id")
                         or AUSENTE),
                     str(linha.get("fonte_tipo") or ""))
            contagem[chave] = contagem.get(chave, 0) + 1
    if not contagem:
        return "Nenhuma referência coletada."
    linhas = ["| Fonte | Natureza | Referências |", "|---|---|---|"]
    for (nome, tipo), quantas in sorted(contagem.items()):
        linhas.append(f"| {nome} | {tipo or AUSENTE} | {quantas} |")
    return "\n".join(linhas)


def quadro_resumo(itens: list[dict]) -> str:
    """§31.11 — uma linha por item, com o essencial."""
    if not itens:
        return "Nenhum item nesta pesquisa."
    linhas = [
        "| Item | Descrição | Unid. | Qtd. | Método | Unitário | Total | "
        "Situação |",
        "|---|---|---|---|---|---|---|---|",
    ]
    for item in sorted(itens, key=lambda i: int(i.get("numero") or 0)):
        linhas.append(
            f"| {int(item.get('numero') or 0):02d} "
            f"| {_texto(item.get('descricao'))} "
            f"| {_texto(item.get('unidade'))} "
            f"| {_numero(item.get('quantidade'))} "
            f"| {_texto(item.get('metodo'))} "
            f"| {_moeda(item.get('preco_estimado'))} "
            f"| {_moeda(item.get('preco_total'))} "
            f"| {_texto(item.get('estado'))} |")
    return "\n".join(linhas)


def _analise_do_item(item: dict,
                     referencias: dict[str, list[dict]]) -> list[str]:
    """
    §31.12 a 17 para um item: referências usadas, descartadas com motivo,
    memória de cálculo, unitário e total.
    """
    numero = int(item.get("numero") or 0)
    lista = referencias.get(str(item.get("id")), [])
    selecionadas = [r for r in lista if str(r.get("status")) == "selected"]
    descartadas = [r for r in lista if str(r.get("status")) != "selected"]

    saida = [f"### Item {numero:02d} — {_texto(item.get('descricao'))}", ""]
    saida.append(f"- **Unidade:** {_texto(item.get('unidade'))}")
    saida.append(f"- **Quantidade:** {_numero(item.get('quantidade'))}")
    if item.get("codigo"):
        saida.append(f"- **Catálogo:** {_texto(item.get('tipo_catalogo'))} "
                     f"{_texto(item.get('codigo'))}")
    else:
        saida.append("- **Catálogo:** não informado — a busca foi feita por "
                     "descrição")
    saida.append(f"- **Situação:** {_texto(item.get('estado'))}")
    saida.append("")

    memoria = item.get("estatisticas") or {}
    estatisticas = memoria.get("estatisticas") or {}
    if estatisticas:
        saida.append("**Estatística da cesta**")
        saida.append("")
        saida.append("| Referências | Menor | Média | Mediana | Maior | CV |")
        saida.append("|---|---|---|---|---|---|")
        saida.append(
            f"| {estatisticas.get('quantidade', 0)} "
            f"| {_moeda(estatisticas.get('menor'))} "
            f"| {_moeda(estatisticas.get('media'))} "
            f"| {_moeda(estatisticas.get('mediana'))} "
            f"| {_moeda(estatisticas.get('maior'))} "
            f"| {_numero(estatisticas.get('coeficiente_variacao'))} |")
        saida.append("")

    saida.append(f"**Referências selecionadas ({len(selecionadas)})** — "
                 "§31.13")
    saida.append("")
    saida.append(_tabela_de_referencias(selecionadas))
    saida.append("")

    # §31.14 — o que ficou de fora NUNCA some do relatório.
    saida.append(f"**Referências desconsideradas ({len(descartadas)})** — "
                 "§31.14")
    saida.append("")
    if descartadas:
        saida.append(_tabela_de_referencias(descartadas, com_motivo=True))
    else:
        saida.append("Nenhuma referência foi desconsiderada neste item.")
    saida.append("")

    saida.append("**Memória de cálculo** — §31.15")
    saida.append("")
    justificativa = str(item.get("justificativa") or "").strip()
    if justificativa:
        for linha in justificativa.splitlines():
            if linha.strip():
                saida.append(f"- {linha.strip()}")
    else:
        saida.append(f"- {AUSENTE}")
    saida.append("")

    anomalias = memoria.get("anomalias") or []
    if anomalias:
        saida.append("**Candidatos discrepantes sinalizados**")
        saida.append("")
        for anomalia in anomalias:
            saida.append(
                f"- {_moeda(anomalia.get('valor'))} — "
                f"{_texto(anomalia.get('motivo'))} "
                f"(critério {_texto(anomalia.get('criterio'))}). "
                "Sinalizado para revisão; não excluído automaticamente.")
        saida.append("")

    saida.append(f"- **Preço estimado unitário (§31.16):** "
                 f"{_moeda(item.get('preco_estimado'))}")
    saida.append(f"- **Valor total do item (§31.17):** "
                 f"{_moeda(item.get('preco_total'))}")
    saida.append("")
    return saida


def _tabela_de_referencias(linhas: list[dict],
                           com_motivo: bool = False) -> str:
    if not linhas:
        return "Nenhuma."
    cabecalho = ("| Fonte | Órgão | UF | Data | Qtd. | Unid. | Unitário | "
                 "Compat. |")
    separador = "|---|---|---|---|---|---|---|---|"
    if com_motivo:
        cabecalho = cabecalho[:-1] + " Situação e motivo |"
        separador += "---|"
    saida = [cabecalho, separador]
    for linha in linhas:
        score = _decimal(linha.get("score"))
        compat = f"{score:.0%}" if score is not None else AUSENTE
        celulas = [
            _texto(linha.get("fonte_nome") or linha.get("fonte_id")),
            _texto(linha.get("orgao")),
            _texto(linha.get("uf")),
            _data(linha.get("data_resultado") or linha.get("data_compra")),
            _numero(linha.get("quantidade_original")),
            _texto(linha.get("unidade_normalizada")
                   or linha.get("unidade_original")),
            _moeda(linha.get("valor_unitario_normalizado")
                   or linha.get("valor_unitario_original")),
            compat,
        ]
        if com_motivo:
            status = _ROTULO_STATUS.get(str(linha.get("status")),
                                        str(linha.get("status") or ""))
            motivos = "; ".join(str(m) for m in (linha.get("motivos") or []))
            celulas.append(f"{status}" + (f" — {motivos}" if motivos else ""))
        saida.append("| " + " | ".join(celulas) + " |")
    return "\n".join(saida)


def _alertas(itens: list[dict], referencias: dict[str, list[dict]],
             perfil) -> list[str]:
    """
    §31.19 — o que o leitor precisa saber antes de usar este preço.

    Escrito em linguagem de RESSALVA, nunca de conclusão jurídica: uma
    fórmula estatística sinaliza dispersão, não ilegalidade.
    """
    saida: list[str] = []
    incompletos = [i for i in itens
                   if str(i.get("estado")) == EstadoItem.INCOMPLETO.value]
    se_erro = [i for i in itens
               if str(i.get("estado")) == EstadoItem.ERRO.value]
    por_revisar = [i for i in itens
                   if str(i.get("estado")) == EstadoItem.EM_REVISAO.value]

    if incompletos:
        numeros = ", ".join(str(int(i.get("numero") or 0))
                            for i in incompletos)
        saida.append(
            f"- **Itens sem preço formado ({len(incompletos)}):** {numeros}. "
            f"A amostra não alcançou o mínimo de {perfil.minimo_referencias} "
            "referências defensáveis exigido pelo perfil, e nenhum preço "
            "foi fabricado para completar a cesta.")
    if por_revisar:
        saida.append(
            f"- **Itens ainda em revisão ({len(por_revisar)}):** o preço "
            "calculado não passou por confirmação humana.")
    if se_erro:
        saida.append(
            f"- **Itens com falha técnica ({len(se_erro)}):** a consulta às "
            "fontes não se completou; a pesquisa desses itens precisa ser "
            "repetida.")

    discrepantes = 0
    manuais = 0
    for item in itens:
        memoria = item.get("estatisticas") or {}
        discrepantes += len(memoria.get("anomalias") or [])
    for lista in referencias.values():
        manuais += sum(1 for r in lista
                       if str(r.get("status")) == "manual_review")
    if discrepantes:
        saida.append(
            f"- **Candidatos discrepantes sinalizados ({discrepantes}):** "
            "valores distantes da mediana da própria amostra. A "
            "sinalização é estatística e pede revisão — não afirma "
            "inexequibilidade nem irregularidade.")
    if manuais:
        saida.append(
            f"- **Referências abaixo do piso de comparabilidade "
            f"({manuais}):** não entraram na cesta automaticamente e "
            "permanecem disponíveis para inclusão manual justificada.")

    ocorrencias: list[str] = []
    for item in itens:
        for ocorrencia in (item.get("ocorrencias") or []):
            if ocorrencia not in ocorrencias:
                ocorrencias.append(str(ocorrencia))
    for ocorrencia in ocorrencias[:10]:
        saida.append(f"- **Ocorrência de fonte:** {ocorrencia}")

    if not saida:
        saida.append("- Nenhuma ressalva registrada nesta pesquisa.")
    return saida


# ---------------------------------------------------------------------------
# §32 — relatório resumido
# ---------------------------------------------------------------------------
PRECOS_EM_DESTAQUE = 3

# No QUADRO compacto a ausência vira travessão, e não "(não informado)".
# Não é inconsistência com o resto do módulo: numa tabela de 210 linhas o
# texto longo repetido em três colunas alarga a tabela até estourar a
# largura útil da página — o defeito de geometria que a Fase 2.1 gastou
# um ciclo inteiro para fechar. Na PROSA e no relatório completo, onde
# não há esse custo, a ausência continua escrita por extenso.
VAZIO_NO_QUADRO = "—" 


def resumido(pesquisa: dict, itens: list[dict],
             referencias: dict[str, list[dict]],
             *, emitido_em: datetime | None = None) -> str:
    """
    A versão compacta do §32: uma linha por item, com os três primeiros
    preços da cesta em destaque e o restante contado.

    Três não é número mágico — é a regra dos três preços da própria Lei,
    e mostrá-los explicitamente é o que permite conferir o cálculo de
    cabeça.
    """
    momento = emitido_em or datetime.now(timezone.utc)
    perfil = obter_perfil(pesquisa.get("perfil_normativo"))
    partes: list[str] = []
    ad = partes.append

    ad("# PESQUISA DE PREÇOS — QUADRO RESUMIDO")
    ad("")
    ad(f"**{_texto(pesquisa.get('nome'))}** · revisão "
       f"{int(pesquisa.get('versao') or 1)} · data-base "
       f"{_data(pesquisa.get('data_base'))}")
    ad("")

    cabecalho = ["Item", "Descrição", "Qtd."]
    cabecalho += [f"Preço {n}" for n in range(1, PRECOS_EM_DESTAQUE + 1)]
    cabecalho += ["Outras", "Média", "Mediana", "Método", "Unitário", "Total"]
    ad("| " + " | ".join(cabecalho) + " |")
    ad("|" + "---|" * len(cabecalho))

    for item in sorted(itens, key=lambda i: int(i.get("numero") or 0)):
        lista = [r for r in referencias.get(str(item.get("id")), [])
                 if str(r.get("status")) == "selected"]
        valores = [
            _decimal(r.get("valor_unitario_normalizado")
                     or r.get("valor_unitario_original"))
            for r in lista]
        valores = sorted(v for v in valores if v is not None)
        destaque = [_moeda(v) for v in valores[:PRECOS_EM_DESTAQUE]]
        destaque += [VAZIO_NO_QUADRO] * (PRECOS_EM_DESTAQUE - len(destaque))
        restantes = max(0, len(valores) - PRECOS_EM_DESTAQUE)

        memoria = (item.get("estatisticas") or {}).get("estatisticas") or {}
        celulas = [
            f"{int(item.get('numero') or 0):02d}",
            _texto(item.get("descricao")),
            _numero(item.get("quantidade")),
            *destaque,
            str(restantes),
            _compacto(memoria.get("media")),
            _compacto(memoria.get("mediana")),
            str(item.get("metodo") or VAZIO_NO_QUADRO),
            _compacto(item.get("preco_estimado")),
            _compacto(item.get("preco_total")),
        ]
        ad("| " + " | ".join(celulas) + " |")
    ad("")

    concluidos = [i for i in itens
                  if str(i.get("estado")) == EstadoItem.COMPLETO.value]
    total = sum((_decimal(i.get("preco_total")) or Decimal("0"))
                for i in concluidos)
    ad(f"**Valor global {_plural_concluidos(len(concluidos))}: "
       f"{_moeda(total)}**")
    ad("")

    ad("## Metodologia e fontes")
    ad("")
    ad(f"- **Perfil normativo:** {perfil.nome} ({perfil.base_legal})")
    ad("- **Formação do preço:** por comparabilidade e prioridade "
       "normativa da fonte, nunca por menor preço. Item sem amostra "
       "defensável fica incompleto — nenhum preço é fabricado.")
    ad(_tabela_de_fontes(referencias))
    ad("")
    ad(f"Emitido em {momento.strftime('%d/%m/%Y às %H:%M')} (UTC). "
       f"Identificador da versão: "
       f"`{identificador_da_versao(pesquisa, itens, referencias)[:16]}…`")
    ad("")
    ad("Este quadro é um resumo. A memória de cálculo, as referências "
       "desconsideradas e os motivos de cada exclusão estão no relatório "
       "completo.")
    return "\n".join(partes)


# ---------------------------------------------------------------------------
# §33 — memória analítica em XLSX
# ---------------------------------------------------------------------------
def xlsx_analitico(pesquisa: dict, itens: list[dict],
                   referencias: dict[str, list[dict]]) -> bytes:
    """
    A planilha da memória analítica: uma aba de itens e uma de
    referências, **incluindo as desconsideradas**.

    É aqui que a pesquisa de 210 itens fica manejável: 6.300 linhas numa
    aba se filtram e se somam; as mesmas 6.300 num PDF ninguém confere.
    O relatório completo continua trazendo tudo — a planilha não o
    substitui, dá outro jeito de ler o mesmo conteúdo.
    """
    from openpyxl import Workbook

    wb = Workbook()

    aba = wb.active
    aba.title = "Itens"
    aba.append(["Item", "Código", "Catálogo", "Descrição", "Unidade",
                "Quantidade", "Situação", "Método", "Preço unitário",
                "Valor total", "Referências", "Na cesta"])
    for item in sorted(itens, key=lambda i: int(i.get("numero") or 0)):
        lista = referencias.get(str(item.get("id")), [])
        aba.append([
            int(item.get("numero") or 0),
            str(item.get("codigo") or ""),
            str(item.get("tipo_catalogo") or ""),
            str(item.get("descricao") or ""),
            str(item.get("unidade") or ""),
            _float(item.get("quantidade")),
            str(item.get("estado") or ""),
            str(item.get("metodo") or ""),
            _float(item.get("preco_estimado")),
            _float(item.get("preco_total")),
            len(lista),
            sum(1 for r in lista if str(r.get("status")) == "selected"),
        ])
    for coluna, largura in zip("ABCDEFGHIJKL",
                              (6, 12, 10, 48, 10, 12, 12, 10, 14, 14, 12, 10)):
        aba.column_dimensions[coluna].width = largura

    detalhe = wb.create_sheet("Referências")
    detalhe.append(["Item", "Situação", "Fonte", "Natureza", "Órgão", "UF",
                    "Município", "Fornecedor", "Marca", "Data",
                    "Descrição na fonte", "Unid. original", "Qtd.",
                    "Preço original", "Unid. normalizada",
                    "Preço normalizado", "Comparabilidade", "Identidade",
                    "Circunstâncias", "Identificador externo",
                    "Impressão da evidência", "Motivos"])
    for item in sorted(itens, key=lambda i: int(i.get("numero") or 0)):
        numero = int(item.get("numero") or 0)
        for linha in referencias.get(str(item.get("id")), []):
            detalhe.append([
                numero,
                _ROTULO_STATUS.get(str(linha.get("status")),
                                   str(linha.get("status") or "")),
                str(linha.get("fonte_nome") or linha.get("fonte_id") or ""),
                str(linha.get("fonte_tipo") or ""),
                str(linha.get("orgao") or ""),
                str(linha.get("uf") or ""),
                str(linha.get("municipio") or ""),
                str(linha.get("fornecedor") or ""),
                str(linha.get("marca") or ""),
                str(linha.get("data_resultado")
                    or linha.get("data_compra") or ""),
                str(linha.get("descricao_original") or ""),
                str(linha.get("unidade_original") or ""),
                _float(linha.get("quantidade_original")),
                _float(linha.get("valor_unitario_original")),
                str(linha.get("unidade_normalizada") or ""),
                _float(linha.get("valor_unitario_normalizado")),
                _float(linha.get("score")),
                _float(linha.get("identidade")),
                _float(linha.get("circunstancias")),
                str(linha.get("id_externo") or ""),
                str(linha.get("raw_hash") or ""),
                "; ".join(str(m) for m in (linha.get("motivos") or [])),
            ])
    for coluna, largura in zip(
            ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"],
            (6, 22, 22, 20, 28, 6, 18, 28, 14, 12, 48)):
        detalhe.column_dimensions[coluna].width = largura

    ficha = wb.create_sheet("Identificação")
    ficha.append(["Campo", "Valor"])
    for rotulo, valor in (
            ("Pesquisa", pesquisa.get("nome")),
            ("Identificador", pesquisa.get("id")),
            ("Revisão", pesquisa.get("versao")),
            ("Perfil normativo", pesquisa.get("perfil_normativo")),
            ("Data-base", pesquisa.get("data_base")),
            ("Situação", pesquisa.get("estado")),
            ("Identificador da versão",
             identificador_da_versao(pesquisa, itens, referencias)),
    ):
        ficha.append([rotulo, str(valor if valor is not None else "")])
    ficha.column_dimensions["A"].width = 24
    ficha.column_dimensions["B"].width = 70

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _compacto(valor) -> str:
    """Moeda para o quadro compacto: ausência vira travessão."""
    numero = _decimal(valor)
    return (VAZIO_NO_QUADRO if numero is None
            else planilha.formatar_moeda(float(numero)))


def _plural_concluidos(quantos: int) -> str:
    """
    "de 1 item concluído" / "de 3 itens concluídos".

    Concordância importa num documento que vai para processo
    administrativo: "dos 1 item(ns)" denuncia texto gerado por máquina e
    tira credibilidade do que está certo no resto da página.
    """
    if quantos == 1:
        return "de 1 item concluído"
    return f"dos {quantos} itens concluídos"


def _float(valor) -> float | None:
    """
    `None` vira célula VAZIA, não zero.

    Zero numa planilha de preços é um preço; vazio é a ausência dele. A
    diferença muda a soma que alguém vai fazer na coluna.
    """
    numero = _decimal(valor)
    return None if numero is None else float(numero)


def nome_do_arquivo(pesquisa: dict, sufixo: str, extensao: str) -> str:
    """Nome estável e legível para os arquivos exportados."""
    bruto = str(pesquisa.get("nome") or "pesquisa").strip()
    seguro = "".join(c if c.isalnum() or c in " -_" else "-" for c in bruto)
    seguro = "-".join(seguro.split())[:60] or "pesquisa"
    versao = int(pesquisa.get("versao") or 1)
    return f"{seguro}-rev{versao}-{sufixo}.{extensao}"
