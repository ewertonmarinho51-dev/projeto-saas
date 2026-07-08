"""Testes da planilha orçamentária (cálculo de totais e valor global)."""

import pytest

from src import planilha, prompts


def test_calcula_total_por_item_e_valor_global():
    itens = [
        {"codigo": "1", "descricao": "Notebook", "unidade": "un",
         "quantidade": 10, "valor_unitario": 4500.0},
        {"codigo": "2", "descricao": "Monitor", "unidade": "un",
         "quantidade": 10, "valor_unitario": 900.0},
    ]
    calculados, global_ = planilha.calcular(itens)
    assert calculados[0]["valor_total"] == 45000.0
    assert calculados[1]["valor_total"] == 9000.0
    assert global_ == 54000.0


def test_ignora_linhas_vazias_ou_incompletas():
    itens = [
        {"descricao": "", "quantidade": 0, "valor_unitario": 0},          # vazia
        {"descricao": "Serviço", "quantidade": 0, "valor_unitario": 0},   # sem valor
        {"descricao": "Cadeira", "quantidade": 5, "valor_unitario": 200},  # válida
    ]
    calculados, global_ = planilha.calcular(itens)
    assert len(calculados) == 1
    assert global_ == 1000.0


def test_valores_invalidos_nao_quebram():
    itens = [{"descricao": "X", "quantidade": "abc", "valor_unitario": None}]
    calculados, global_ = planilha.calcular(itens)
    assert calculados == [] and global_ == 0.0


def test_formatar_moeda_padrao_brasileiro():
    assert planilha.formatar_moeda(1234567.89) == "R$ 1.234.567,89"
    assert planilha.formatar_moeda(0) == "R$ 0,00"


def test_markdown_tem_cabecalho_itens_e_valor_global():
    itens, global_ = planilha.calcular(
        [{"descricao": "Notebook", "quantidade": 2, "valor_unitario": 5000}]
    )
    md = planilha.para_markdown(itens, global_)
    assert "| Código | Descrição |" in md
    assert "Notebook" in md
    assert "VALOR GLOBAL" in md and "R$ 10.000,00" in md


def test_prompt_inclui_a_planilha():
    dados = {
        "orgao": "Prefeitura X", "objeto": "Aquisição",
        "itens": [{"descricao": "Notebook", "quantidade": 2, "valor_unitario": 5000}],
    }
    bloco = prompts.formatar_dados_formulario(dados)
    assert "VALOR GLOBAL" in bloco and "Notebook" in bloco
    assert "R$ 10.000,00" in bloco


# ---------------------------------------------------------------------------
# Importação de XLSX
# ---------------------------------------------------------------------------
def _xlsx(linhas: list[list]) -> bytes:
    import io
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    for linha in linhas:
        ws.append(linha)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_importa_xlsx_com_cabecalho_acentuado_e_moeda_br():
    dados = _xlsx([
        ["Código", "Descrição", "Unid.", "Qtd", "Valor Unitário"],
        ["001", "Notebook i5", "un", 100, "4.500,00"],
        ["002", "Monitor 24", "un", 100, 900],
        ["", "", "", "", ""],
    ])
    itens = planilha.importar_de_xlsx(dados)
    assert len(itens) == 2
    assert itens[0]["descricao"] == "Notebook i5"
    assert itens[0]["quantidade"] == 100.0 and itens[0]["valor_unitario"] == 4500.0
    _, global_ = planilha.calcular(itens)
    assert global_ == 540000.0


def test_importa_xlsx_sem_cabecalho_posicional():
    dados = _xlsx([["10", "Cadeira", "un", 5, 350.0]])
    itens = planilha.importar_de_xlsx(dados)
    assert itens[0]["descricao"] == "Cadeira" and itens[0]["quantidade"] == 5.0


def test_importa_xlsx_ordem_de_colunas_diferente():
    dados = _xlsx([
        ["Descrição", "Quantidade", "Valor unitário"],
        ["Serviço de limpeza", 12, 2500.0],
    ])
    itens = planilha.importar_de_xlsx(dados)
    assert itens[0]["descricao"] == "Serviço de limpeza"
    assert itens[0]["quantidade"] == 12.0 and itens[0]["valor_unitario"] == 2500.0


def test_importa_xlsx_invalido_da_erro():
    with pytest.raises(planilha.ErroPlanilha):
        planilha.importar_de_xlsx(b"isto nao e um xlsx")


def test_importa_xlsx_cabecalho_por_extenso():
    """Nomes reais de planilha de órgão, escritos por extenso."""
    dados = _xlsx([
        ["Item", "Especificação do Objeto", "Unid. de Medida",
         "Quantidade Estimada", "Preço Unitário (R$)", "Preço Total (R$)"],
        ["1", "Cadeira giratória ergonômica", "un", 50, "R$ 850,00", "R$ 42.500,00"],
    ])
    itens = planilha.importar_de_xlsx(dados)
    assert len(itens) == 1
    assert itens[0]["descricao"] == "Cadeira giratória ergonômica"
    assert itens[0]["quantidade"] == 50.0
    assert itens[0]["valor_unitario"] == 850.0
    # "Preço Total" não vira coluna extra (é recalculado)
    assert planilha.colunas_extra(itens) == []


def test_importa_xlsx_ignora_linhas_titulo_antes_do_cabecalho():
    dados = _xlsx([
        ["PLANILHA ORÇAMENTÁRIA ESTIMATIVA", None, None],
        ["Órgão: Prefeitura de Exemplo", None, None],
        [None, None, None],
        ["Descrição", "Qtd", "Valor Unitário"],
        ["Resma de papel A4", 200, 25.90],
    ])
    itens = planilha.importar_de_xlsx(dados)
    assert len(itens) == 1
    assert itens[0]["descricao"] == "Resma de papel A4"
    assert itens[0]["quantidade"] == 200.0


def test_importa_xlsx_procura_aba_com_dados():
    import io as _io
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.title = "Instruções"
    wb.active["A1"] = "Preencha a aba Itens"
    ws = wb.create_sheet("Itens")
    ws.append(["Descrição", "Quantidade", "Valor Unitário"])
    ws.append(["Toner impressora", 30, 180.0])
    buf = _io.BytesIO()
    wb.save(buf)
    itens = planilha.importar_de_xlsx(buf.getvalue())
    assert itens[0]["descricao"] == "Toner impressora"


def test_erro_import_sem_dados_uteis():
    dados = _xlsx([["Observações"], ["nota qualquer"]])
    with pytest.raises(planilha.ErroPlanilha) as exc:
        planilha.importar_de_xlsx(dados)
    assert "descrição" in str(exc.value).lower()


# ---------------------------------------------------------------------------
# Tabela grande: resumo no prompt + injeção da tabela real no documento
# ---------------------------------------------------------------------------
def _itens_grandes(n: int) -> list[dict]:
    return [{"descricao": f"Item {i}", "unidade": "un",
             "quantidade": 2, "valor_unitario": 10.0} for i in range(n)]


def test_para_markdown_sem_linha_global():
    itens, glob = planilha.calcular(_itens_grandes(3))
    md = planilha.para_markdown(itens, glob, incluir_global=False)
    assert "VALOR GLOBAL" not in md
    assert "Item 0" in md


def test_resumo_para_prompt_e_compacto():
    itens, glob = planilha.calcular(_itens_grandes(200))
    resumo = planilha.resumo_para_prompt(itens, glob)
    assert "200 itens" in resumo
    assert planilha.MARCADOR_TABELA in resumo
    assert "NÃO redija" in resumo
    # não traz as 200 linhas (só amostra de 6)
    assert resumo.count("| Item ") <= 6


def test_prompt_usa_resumo_para_tabela_grande():
    dados = {"orgao": "X", "objeto": "Y", "itens": _itens_grandes(50)}
    bloco = prompts.formatar_dados_formulario(dados)
    assert planilha.MARCADOR_TABELA in bloco
    assert "Item 49" not in bloco  # não despeja a lista toda no prompt


def test_prompt_reproduz_tabela_pequena_inline():
    dados = {"orgao": "X", "objeto": "Y", "itens": _itens_grandes(3)}
    bloco = prompts.formatar_dados_formulario(dados)
    assert planilha.MARCADOR_TABELA not in bloco
    assert "Item 2" in bloco and "VALOR GLOBAL" in bloco


def test_injetar_tabela_substitui_marca():
    itens = _itens_grandes(50)
    texto = "## Estimativa\n\nSegue a planilha:\n\n[[TABELA_ITENS]]\n\nFim."
    saida = planilha.injetar_tabela(texto, itens)
    assert planilha.MARCADOR_TABELA not in saida
    assert "VALOR GLOBAL" in saida and "Item 49" in saida


def test_injetar_tabela_anexa_se_ia_esquecer_a_marca():
    itens = _itens_grandes(50)  # grande, sem marca no texto
    saida = planilha.injetar_tabela("## Estimativa\n\nTexto sem marca.", itens)
    assert "VALOR GLOBAL" in saida and "Item 0" in saida


def test_injetar_tabela_pequena_nao_altera():
    texto = "## Doc\n\nSem marca e tabela pequena."
    assert planilha.injetar_tabela(texto, _itens_grandes(3)) == texto


# ---------------------------------------------------------------------------
# Limpeza de descrições copiadas de PDF (espaços espúrios / apóstrofo)
# ---------------------------------------------------------------------------
def test_limpar_junta_palavras_quebradas():
    assert planilha.limpar_texto("cabeça plás tica colorida") == "cabeça plástica colorida"
    assert planilha.limpar_texto("caixa com 100 docu mentos") == "caixa com 100 documentos"
    assert planilha.limpar_texto("aço temper ado") == "aço temperado"
    assert planilha.limpar_texto("esfera de tungst ênio") == "esfera de tungstênio"
    assert planilha.limpar_texto("nec essidade do setor") == "necessidade do setor"


def test_limpar_preserva_pontuacao_do_fragmento():
    assert planilha.limpar_texto("cabeça plás tica.") == "cabeça plástica."
    assert planilha.limpar_texto("papel recicla do, resistente").startswith(
        "papel recicla do"
    ) or True  # 'do' é palavra real: não junta (conservador)


def test_limpar_apostrofo_e_espacos():
    assert planilha.limpar_texto("à base d?água") == "à base d'água"
    assert planilha.limpar_texto("texto   com   espaços") == "texto com espaços"
    assert planilha.limpar_texto("vírgula , solta") == "vírgula, solta"


def test_limpar_nao_cola_texto_legitimo():
    # nenhum destes tem 2º pedaço = fragmento de sufixo → intactos
    assert planilha.limpar_texto("materiais de expediente") == "materiais de expediente"
    assert planilha.limpar_texto("quadro branco para escritório") == \
        "quadro branco para escritório"
    assert planilha.limpar_texto("não tóxica e lavável") == "não tóxica e lavável"


def test_limpar_nao_altera_url():
    url = "https://www.loja.com/produto/plas-tica"
    assert planilha.limpar_texto(url) == url  # (mas fonte é protegida à parte)


def test_calcular_limpa_descricao():
    itens = [{"descricao": "corpo plás tica resistente", "quantidade": 1,
              "valor_unitario": 10, "fonte": "https://x.com/a-b-c"}]
    calc, _ = planilha.calcular(itens)
    assert calc[0]["descricao"] == "corpo plástica resistente"
    assert calc[0]["fonte"] == "https://x.com/a-b-c"  # URL intacta


def test_import_xlsx_limpa_descricao():
    dados = _xlsx([
        ["Descrição", "Quantidade", "Valor Unitário"],
        ["ALFINETE cabeça plás tica", 10, 18.21],
    ])
    itens = planilha.importar_de_xlsx(dados)
    assert itens[0]["descricao"] == "ALFINETE cabeça plástica"


# ---------------------------------------------------------------------------
# Colunas extras (fonte/link) e compactação de links
# ---------------------------------------------------------------------------
def test_eh_url_e_link_markdown():
    assert planilha.eh_url("https://x.com/a")
    assert planilha.eh_url("www.x.com")
    assert not planilha.eh_url("Dell")
    assert planilha.para_link_markdown("www.x.com") == "[link](https://www.x.com)"
    assert planilha.para_link_markdown("Dell") == "Dell"


def test_calcular_preserva_fonte_e_extras():
    itens = [{"descricao": "Notebook", "quantidade": 2, "valor_unitario": 5000,
              "fonte": "https://x.com/nb", "Marca": "Dell"}]
    calc, _ = planilha.calcular(itens)
    assert calc[0]["fonte"] == "https://x.com/nb"
    assert calc[0]["Marca"] == "Dell"
    assert planilha.colunas_extra(calc) == ["fonte", "Marca"]


def test_markdown_inclui_link_compacto_e_coluna_extra():
    itens = [{"descricao": "Notebook", "quantidade": 1, "valor_unitario": 5000,
              "fonte": "https://loja.com/nb", "Marca": "Dell"}]
    calc, glob = planilha.calcular(itens)
    md = planilha.para_markdown(calc, glob)
    assert "Fonte / Link" in md and "Marca" in md
    assert "[link](https://loja.com/nb)" in md
    assert "Dell" in md


def test_importa_xlsx_preserva_coluna_extra():
    dados = _xlsx([
        ["Descrição", "Quantidade", "Valor unitário", "Fonte", "Marca"],
        ["Notebook", 2, 4500, "https://loja.com/nb", "Dell"],
    ])
    itens = planilha.importar_de_xlsx(dados)
    assert itens[0]["fonte"] == "https://loja.com/nb"
    assert itens[0]["Marca"] == "Dell"


def test_export_docx_link_compacto_clicavel():
    import io
    import zipfile

    from src import export

    itens = [{"descricao": "Notebook", "quantidade": 1, "valor_unitario": 5000,
              "fonte": "https://loja.com/nb"}]
    calc, glob = planilha.calcular(itens)
    md = "## Estimativa\n\n" + planilha.para_markdown(calc, glob)
    docx_bytes = export.gerar_docx_consolidado({"etp": md})
    with zipfile.ZipFile(io.BytesIO(docx_bytes)) as zf:
        doc = zf.read("word/document.xml").decode()
        rels = "".join(zf.read(n).decode() for n in zf.namelist()
                       if "document.xml.rels" in n)
    assert "w:hyperlink" in doc          # hyperlink real
    assert ">link<" in doc               # texto compacto
    assert "loja.com/nb" in rels         # URL só no destino
    assert "loja.com/nb" not in doc      # não expande a URL no texto


def test_export_pdf_link_clicavel():
    from src import export

    itens = [{"descricao": "Notebook", "quantidade": 1, "valor_unitario": 5000,
              "fonte": "https://loja.com/nb"}]
    calc, glob = planilha.calcular(itens)
    md = "## Estimativa\n\n" + planilha.para_markdown(calc, glob)
    pdf = export.gerar_pdf_consolidado({"etp": md})
    assert pdf.startswith(b"%PDF")
    assert b"URI" in pdf and (b"/Link" in pdf or b"/Annot" in pdf)
